#!/usr/bin/env python3
"""
Convert the curated source Excel into src/data/data.json.

Re-run this whenever a new Excel drop arrives. Two modes:

  # a fresh drop that ALSO ships a new image folder -> match + copy images:
  python3 scripts/convert_data.py \
      --xlsx "/path/to/_MASTERFILE all sources sorted.xlsx" \
      --images "/path/to/data send folder"

  # a data-only drop (no new images) -> keep the images already in
  # public/img/recipeImages and re-attach them to records by id:
  python3 scripts/convert_data.py --carry-images

Defaults point at the _MASTERFILE in ~/Projects/_privat.

It does three things:
  1. reads every recipe row from the MASTERFILE sheet (carrying forward
     blank file/author/country),
  2. parses the free-text ingredient column into {ingredient, amount, unit} best-effort,
  3. either copies + matches a source image folder into public/img/recipeImages,
     OR (with --carry-images) re-uses the image URLs already in data.json,
     matched back to records by id.

A report is printed at the end (counts, unparsed ranges, unmatched/carried images).
"""

import argparse
import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path

import openpyxl

PROJECT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path.home() / "Projects/_privat/_MASTERFILE all sources sorted[22].xlsx"
DEFAULT_IMAGES = Path.home() / "Downloads/260611 data send"
OUT_JSON = PROJECT / "src/data/data.json"
IMG_DIR = PROJECT / "public/img/recipeImages"
IMG_URL_PREFIX = "/img/recipeImages"

# Column letters on the MASTERFILE sheet (label row is row 2, data from row 4).
# Note vs. earlier drops: a new column L ("TITLE Website") was inserted to hold
# the curated, human-facing display title, so everything from M onward shifted
# one to the right. This website title is what we now use for the record `name`
# (falling back to the raw translated/original title). Column A ("YEAR sort") is
# the pre-cleaned integer date; G ("Score") is an internal field and is
# intentionally NOT mapped.
COL = {
    "year": "A", "year_real": "B", "file": "C", "author": "D", "country": "E",
    "type": "F", "source": "H", "bibl": "I", "page": "J", "abbr": "K",
    "title_website": "L",
    "title_tr": "M", "text_tr": "N", "ingredients": "O", "transl_kind": "P",
    "comment_tr": "Q", "title_orig": "R", "text_orig": "S", "comment_text": "T",
    "link": "U", "notes": "V", "further_bibl": "W",
}

# Units we recognise at the start of an ingredient line. Longer/multiword first.
# "d" is an old weight/measure abbreviation used in the accounts (e.g. "6 d. pitch").
UNITS = [
    "quarter of a pound", "hände voll", "hand voll", "handfuls", "handfull",
    "handful", "hundredweight", "pfund",
    "loth", "lot", "quentchen", "quintlein", "quentl", "unze", "unzen", "spoonfuls", "spoonful",
    "measures", "measure", "maß", "mass", "maas", "eimer", "nössel",
    "seidel", "gran", "theil", "teil", "libbra", "libbre", "libre", "oncia",
    "once", "pound", "pounds", "stone", "stones", "gallon", "gallons", "pint",
    "pints", "quart", "ounce", "ounces", "oz", "lb", "lbs", "jug", "jugs",
    "bushels", "bushel", "parts", "part", "bits", "bit", "eh", "g", "kg", "ml", "l", "d",
]
# Normalised lookup used to recognise a bracketed unit gloss ("[Loth]", "[Pfund]").
UNIT_SET = {u.lower() for u in UNITS}
# Build a regex alternation, escaping, with optional trailing dot.
_UNIT_RE = "|".join(sorted((re.escape(u) for u in UNITS), key=len, reverse=True))
# Leading numeric quantity: 19  6,5  1/2  1½  ½  and ranges 2-3, 1/16 - 1/8 ...
_NUM = r"\d+(?:[.,]\d+)?"          # 19  6,5
_FRAC = r"\d+\s*/\s*\d+"          # 1/2  1/16
_UNI = r"[½¼¾⅓⅔⅛]"               # unicode fractions
# a single quantity value: a mixed number ("1 1/2"), a fraction, a number
# (optionally + a unicode fraction like "1½"), or a bare unicode fraction. The
# mixed form comes first so "1 1/2" is taken whole; the bare fraction next so
# "1/16" isn't cut to "1".
_VALUE = rf"(?:{_NUM}\s+{_FRAC}|{_FRAC}|{_NUM}{_UNI}?|{_UNI})"
# a quantity is one value, optionally a dash-range to a second value ("2-3",
# "1/16 - 1/8") — so the whole range is the amount, not part of the ingredient.
_QTY = rf"{_VALUE}(?:\s*[-–]\s*{_VALUE})?"
# A leading quantity, anchored.
QTY_RE = re.compile(rf"^(?:{_QTY})", re.IGNORECASE)
# A unit, but ONLY as a whole token (followed by space, dot, or end) — so we never
# slice a unit off the front of a word ("leadwhite" -> l|eadwhite, "stone dust" ->
# stone|dust). A unit is also only honoured when it follows a quantity (see
# parse_ingredients): a bare "stone"/"glass"/"lime" is an ingredient, not a measure.
UNIT_RE = re.compile(rf"(?:{_UNIT_RE})(?=[\s.]|$)", re.IGNORECASE)
# A TRAILING quantity: "resin 1 part", "wax 2 loth", "stone dust 3 pounds",
# "Resin 7 - 8 parts". When a line has no LEADING quantity, the measure is often
# written after the ingredient (name + amount + unit). We only treat the tail as
# a measure when a real unit token closes the line, so a plain name that merely
# ends in a number ("no. 2") is left untouched.
TRAIL_RE = re.compile(rf"\s+({_QTY})\s+({_UNIT_RE})\.?$", re.IGNORECASE)
# Spelled-out quantities written before a unit: "a bushel …", "two handfuls …".
WORDED = {"a": "1", "an": "1", "one": "1", "two": "2", "three": "3", "four": "4",
          "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
          "ten": "10", "half": "1/2"}
_WORDED_RE = "|".join(sorted(WORDED, key=len, reverse=True))


def strip_diacritics(s: str) -> str:
    # Keep German umlaut transliteration consistent with existing files (ö->oe ...).
    s = unicodedata.normalize("NFC", s)  # macOS stores filenames as NFD (decomposed)
    table = {"ö": "oe", "ü": "ue", "ä": "ae", "ß": "ss", "Ö": "Oe", "Ü": "Ue", "Ä": "Ae",
             "ø": "o", "Ø": "O", "å": "a", "Å": "A", "æ": "ae", "é": "e", "è": "e"}
    out = "".join(table.get(c, c) for c in s)
    out = unicodedata.normalize("NFKD", out)
    return "".join(c for c in out if not unicodedata.combining(c))


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_year(raw, prev):
    """Clean integer date from column A ("YEAR sort"). This column is already
    cleaned by the supplier (no ranges, no "c."), so we just coerce to int and
    fall back to the first 3-4 digit run / the carried-forward year."""
    if raw is None:
        return prev
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip()
    if not s or s.lower() == "same":
        return prev
    m = re.search(r"\d{3,4}", s)
    return int(m.group(0)) if m else prev


def display_date(raw, date_int):
    """The human-facing date from column B ("YEAR real"), which may carry the
    nuance the sort year drops: ranges ("1301-1302"), approximations ("c.1400"),
    or a republished-vs-original note ("1904 (1880s)"). Returns None when it is
    just the plain sort year again (so the detail view falls back to `date`)."""
    if raw is None:
        return None
    s = re.sub(r"\s+", " ", str(raw)).strip()
    if not s:
        return None
    if re.fullmatch(r"\d{3,4}", s) and date_int is not None and int(s) == date_int:
        return None
    return s


def parse_page(raw):
    """'p.159' -> '159';  'p.122-123' -> '122-123'. Kept as a string (some
    pages are ranges), but without the 'p.' prefix — the UI prepends it."""
    if raw is None:
        return ""
    s = re.sub(r"\s+", " ", str(raw).strip())
    return re.sub(r"^[pP]\.\s*", "", s)  # strip one leading "p."


def split_author(full):
    if not full or full.strip().lower() == "unknown":
        return "", "unknown"
    parts = full.split()
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def parse_ingredients(raw):
    """Free text -> [{ingredient, amount, unit}]. Split on newlines (the only safe
    boundary, since '6,5' uses a decimal comma)."""
    if not raw:
        return []
    out = []
    for line in str(raw).split("\n"):
        line = line.strip().strip(";,").strip()
        if not line:
            continue
        # A leading parenthetical that starts with a number is really the quantity
        # written in brackets: "(1 part) Pitch..." -> "1 part Pitch...". A wholly
        # bracketed line unwraps the same way ("(3 stones of onions)" -> the tail
        # is empty, so we just drop the brackets and parse the inside).
        mp = re.match(r"^\(\s*([^)]*)\)\s*(.*)$", line)
        if mp and re.match(r"\s*\d", mp.group(1)):
            line = f"{mp.group(1).strip()} {mp.group(2).strip()}".strip()
        amount, unit, name = "", "", line
        qm = QTY_RE.match(line)
        if qm:
            amount = qm.group(0).strip()
            rest = line[qm.end():].lstrip()
            # A unit only counts right after the quantity, and only as a whole token.
            um = UNIT_RE.match(rest)
            if um:
                unit = um.group(0).rstrip(".").strip()
                # UNIT_RE stops before the unit's own abbreviation dot ("lb." or
                # "lb ."), so the leftover starts with ". " — strip that (and any
                # stray leading punctuation) so we get "wax", not ". wax".
                rest = re.sub(r"^[\s.]+", "", rest[um.end():])
            # A bracketed unit gloss ("[Loth]", "[Pfund]") sits where the unit
            # would be — sometimes duplicating an English unit already parsed
            # ("1 pound [Pfund] resin"). Unwrap it, adopting it as the unit if we
            # don't have one yet, and drop the brackets from the name either way.
            mb = re.match(r"^\[\s*([^\]]+?)\s*\]\s*", rest)
            if mb and mb.group(1).lower() in UNIT_SET:
                if not unit:
                    unit = mb.group(1).strip()
                rest = rest[mb.end():]
            name = rest
        else:
            # A spelled-out quantity before a unit: "a bushel tyldust" (=1),
            # "two handfuls marble" (=2). Only fires right before a real unit, so
            # "a little wax" ("little" is no unit) stays a plain name.
            am = re.match(rf"^({_WORDED_RE})\s+({_UNIT_RE})(?=[\s.])\s+(.+)$",
                          line, re.IGNORECASE)
            if am:
                amount = WORDED[am.group(1).lower()]
                unit = am.group(2).rstrip(".").strip()
                name = am.group(3).strip()
            else:
                # No leading quantity — the measure may trail the ingredient instead
                # ("resin 1 part", "wax 2 loth", "ground chalk 0.25 pound"). Only split
                # off the tail when a real unit closes the line and a name precedes it.
                tm = TRAIL_RE.search(line)
                if tm and line[:tm.start()].strip():
                    amount = re.sub(r"\s*[-–]\s*", "-", tm.group(1).strip())  # "7 - 8"->"7-8"
                    unit = tm.group(2).rstrip(".").strip()
                    name = line[:tm.start()].strip()
        name = name.strip().strip(".").strip()
        # a leftover leading connector from "<qty> <unit> of <ingredient>"
        # ("1 pound of wax" -> "wax") or an alternative line ("or unslacked lime
        # and pig fat" -> "unslacked lime and pig fat") is not part of the name.
        name = re.sub(r"^(?:of|or)\s+", "", name, flags=re.IGNORECASE)
        # a line that is nothing but a connector ("or", "and") is a split artifact
        # from "X or Y" / "X and Y" spread over lines — drop it entirely.
        if name.strip().lower() in {"of", "or", "and", "&", "oder", "und"}:
            continue
        # If stripping the quantity/unit left no name, the line was really just a name.
        if not name:
            amount, unit, name = "", "", line
        out.append({"ingredient": name, "amount": amount, "unit": unit})
    return out


# ---- image matching ------------------------------------------------------
#
# Both record ids ("1844.Tho.04a") and image filenames ("1844.Tho.04a.1.png",
# "Crö_1736_01", "Mox01", "1681_Bal_A", "p.386-387") encode the same thing:
# (author abbreviation, recipe index, optional a/b letter variant). We parse
# both into a common (abbr, idx, letter) key and match on it. Images with no
# index (covers, title pages, lettered plates, page scans) are "source-level"
# and get attached to every recipe of that source.

ABBR_STOP = {"p", "pp", "hq", "shot", "cover", "title", "inner", "page", "pages",
             "tavola", "tav", "fol", "fig", "plate"}


def _norm(s):
    return strip_diacritics(s).lower()


def keys_from_record(rid):
    """Record id -> [(abbr, idx, letter)]. Always a single key (in a list for
    symmetry with filenames, which can expand to ranges)."""
    base = _norm(rid.split("#")[0])
    m = re.match(r"(\d{3,4})\.([a-z]{2,6})\.0*(\d+)([a-z])?", base)
    if not m:
        return []
    return [(m.group(2), int(m.group(3)), m.group(4))]


def parse_filename(name, valid_abbrs):
    """-> dict(abbr, keys=[(abbr,idx,letter)...], page=int|None, source_level=bool).
    `valid_abbrs` is the set of author codes seen in records, used to locate the
    abbreviation robustly and to detect naming inconsistencies."""
    base = _norm(Path(name).stem)
    tokens = re.split(r"[.\s_]+", base)  # keep '-' so we can spot ranges

    year = next((int(t) for t in tokens
                 if re.fullmatch(r"\d{3,4}", t) and 1000 <= int(t) <= 2099), None)

    # locate the author abbreviation (a token, possibly glued to digits)
    abbr = None
    abbr_pos = None
    abbr_like = None  # an abbr-looking token even if not in valid_abbrs (for flagging)
    for i, t in enumerate(tokens):
        mt = re.match(r"^([a-z]{2,6})", t)
        cand = mt.group(1) if mt else None
        if not cand or cand in ABBR_STOP:
            continue
        if abbr_like is None and not re.fullmatch(r"\d{3,4}", t):
            abbr_like = cand
        if cand in valid_abbrs:
            abbr, abbr_pos = cand, i
            break

    # page-named scan? ("p.386-387", "p.46", "p.0")
    pm = re.search(r"\bp\.?\s*(\d+(?:-\d+)?)", base)
    pages = page_numbers(pm.group(1)) if pm else set()

    if abbr is None:
        # A short foreign code + an index token (e.g. 'Rai_1399_01' while records
        # use 'Yor') is a naming inconsistency we must NOT silently map.
        has_index = any(re.fullmatch(r"0*\d{1,3}[a-z]?", t) for t in tokens)
        foreign_indexed = bool(abbr_like and len(abbr_like) <= 4 and has_index)
        return {"abbr": None, "abbr_like": abbr_like, "keys": [], "pages": pages,
                "year": year, "source_level": True, "foreign_indexed": foreign_indexed}

    # find the index token: first token at/after the abbr that carries a number,
    # ignoring the year. Handles 'mox01', '04a', '01a', '02-03', '05a-e'.
    keys = []
    for t in tokens[abbr_pos:]:
        body = t
        m = re.match(rf"^{re.escape(abbr)}(\d.*)$", body)  # 'mox01' -> '01'
        if m:
            body = m.group(1)
        if re.fullmatch(r"\d{3,4}", body):  # a year, not an index
            continue
        # number range: 02-03
        mr = re.fullmatch(r"0*(\d+)-0*(\d+)", body)
        if mr:
            for n in range(int(mr.group(1)), int(mr.group(2)) + 1):
                keys.append((abbr, n, None))
            break
        # letter range: 05a-e
        ml = re.fullmatch(r"0*(\d+)([a-z])-([a-z])", body)
        if ml:
            for c in range(ord(ml.group(2)), ord(ml.group(3)) + 1):
                keys.append((abbr, int(ml.group(1)), chr(c)))
            break
        # plain index with optional single letter: 04, 04a
        ms = re.fullmatch(r"0*(\d+)([a-z])?", body)
        if ms:
            keys.append((abbr, int(ms.group(1)), ms.group(2)))
            break
    source_level = not keys  # e.g. lettered plate 'Bal_A', or only a cover image
    return {"abbr": abbr, "abbr_like": abbr_like, "keys": keys, "pages": pages,
            "year": year, "source_level": source_level, "foreign_indexed": False}


def page_numbers(page_str):
    """'122-123' -> {122, 123}; '395' -> {395}; '' -> set()."""
    out = set()
    nums = [int(n) for n in re.findall(r"\d+", page_str or "")]
    if len(nums) == 2 and 0 <= nums[1] - nums[0] <= 50:
        out.update(range(nums[0], nums[1] + 1))
    else:
        out.update(nums)
    return out


def sanitize_filename(name):
    stem, ext = Path(name).stem, Path(name).suffix
    s = strip_diacritics(stem)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")
    return f"{s}{ext.lower()}"


# ---- canonical image naming ----------------------------------------------
#
# The supplier's filenames mix conventions (dotted "1844.Tho.01", underscored
# "Croe_1736_01"/"1640_Sto_01", year-less "Bor_01a", glued "Mox01", descriptive
# "shot"/"inner cover"/"p.46 hq"). To give the PROJECT one stable scheme that
# survives re-generation and future drops, every image is renamed to
#     <YEAR>.<Abbr>.<tail>
# where YEAR and Abbr come from the matched record's id (ASCII-transliterated,
# so 1736.Crö -> "1736.Croe"), and <tail> is the index / page / plate / label
# normalised out of the source filename.

# Descriptor tokens we drop entirely (noise) or never treat as part of the tail.
_DROP_TOKENS = {"hq", "page", "inner", "kunckel", "kunkel"}


def make_tail(stem, year, abbr):
    """The part after <YEAR>.<Abbr>, normalised from the source filename.
    e.g. ('Croe_1736_02a', 1736, 'Croe') -> '02a';  ('p.46 hq', .., 'Kun') -> 'p46';
         ('inner cover', .., 'Mor') -> 'cover';  ('Mox01', 1703, 'Mox') -> '01'."""
    al = abbr.lower()
    s = strip_diacritics(stem)
    s = re.sub(r"[_\s]+", ".", s)              # unify separators to dots (keep '-')
    out = []
    for t in s.split("."):
        if not t:
            continue
        tl = t.lower()
        if t == str(year) or tl == al or tl in _DROP_TOKENS:
            continue
        # the source's full name spelt out ("Neves" when abbr is "Nev"): drop it
        if t.isalpha() and tl.startswith(al) and len(t) > len(al):
            continue
        # abbr glued to the index ("Mox01" -> "01")
        mg = re.match(rf"^{re.escape(al)}(\d.*)$", tl)
        if mg:
            t = mg.group(1)
        out.append(t)
    if not out:
        return ""
    # a leading page marker glues to its number: ['p','46'] -> 'p46', ['p','IV'] -> 'pIV'
    if out[0].lower() == "p" and len(out) >= 2:
        rest = "".join("." + x for x in out[2:])
        return f"p{out[1]}{rest}"
    # zero-pad a leading bare index ('1' -> '01'; '04a','02-03','05a-e' kept as-is)
    m0 = re.fullmatch(r"(\d+)([a-z].*)?", out[0])
    if m0:
        out[0] = f"{int(m0.group(1)):02d}{m0.group(2) or ''}"
    return ".".join(out)


def collect_images(images_root: Path):
    files = []
    for p in sorted(images_root.rglob("*")):
        if p.is_file() and p.suffix.lower() in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            files.append(p)
    return files


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--images", type=Path, default=DEFAULT_IMAGES)
    ap.add_argument("--no-images", action="store_true", help="skip image copy/match")
    ap.add_argument("--carry-images", action="store_true",
                    help="data-only drop: keep the images already in public/img/"
                         "recipeImages and re-attach them from the existing data.json "
                         "by record id (no source folder needed, IMG_DIR untouched)")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"Excel not found: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ci = openpyxl.utils.column_index_from_string

    def cell(r, key):
        return clean(ws.cell(row=r, column=ci(COL[key])).value)

    records = []
    prev = {"year": None, "file": None, "author": None, "country": None}
    seen_ids = {}
    ranges = []

    for r in range(4, ws.max_row + 1):
        # skip fully empty rows
        if all(cell(r, k) is None for k in COL):
            continue
        # skip rows with no recipe content at all (e.g. a stray editorial note left
        # in the YEAR column): a real recipe has at least an id, a title, body text,
        # or ingredients. Otherwise carry-forward would mint a phantom "(untitled)".
        if not any(cell(r, k) for k in ("abbr", "title_website", "title_tr",
                                        "title_orig", "text_tr", "text_orig",
                                        "ingredients")):
            continue
        # carry-forward fields
        for k in ("file", "author", "country"):
            v = cell(r, k)
            if v and v.lower() != "same":
                prev[k] = v
        date = parse_year(ws.cell(row=r, column=ci(COL["year"])).value, prev["year"])
        prev["year"] = date
        date_display = display_date(ws.cell(row=r, column=ci(COL["year_real"])).value, date)
        if date_display:
            ranges.append((cell(r, "abbr"), date_display))

        src_type = cell(r, "type") or ""
        author = cell(r, "author") or prev["author"]
        first, last = split_author(author)
        # Accounts (and other author-less sources, e.g. the Report) have no known
        # author. Per the supplier's request we surface the place (the FILE column,
        # e.g. "Westminster Palace") in the author slot instead; the date is shown
        # alongside as usual, giving the requested "date and place". Manuals (and the
        # Instruction, which names Richard Voigtel) keep their real author name.
        author_known = bool(author) and author.strip().lower() != "unknown"
        if not author_known:
            first, last = "", (cell(r, "file") or prev["file"] or "")

        abbr = cell(r, "abbr") or f"row{r}"
        rid = abbr
        if rid in seen_ids:
            seen_ids[rid] += 1
            rid = f"{abbr}#{seen_ids[abbr]}"
        else:
            seen_ids[abbr] = 1

        rec = {
            "id": rid,
            "authorFirstName": first,
            "authorLastName": last,
            "date": date,
            "manual": cell(r, "source") or cell(r, "file") or "",
            "page": parse_page(ws.cell(row=r, column=ci(COL["page"])).value),
            "name": cell(r, "title_website") or cell(r, "title_tr")
            or cell(r, "title_orig") or "(untitled)",
            "images": [],
            "ingredients": parse_ingredients(cell(r, "ingredients")),
            "description": cell(r, "text_tr") or "",
            "source": cell(r, "bibl") or "",
            "type": src_type,
            "url": cell(r, "link"),
            # --- detail-view only (not in the table) ---
            "country": cell(r, "country") or prev["country"] or "",
            "site": cell(r, "file") or "",
            "dateDisplay": date_display,
            "originalTitle": cell(r, "title_orig"),
            "originalText": cell(r, "text_orig"),
            "notes": cell(r, "notes"),
            "furtherBibliography": cell(r, "further_bibl"),
            "commentOnText": cell(r, "comment_text"),
        }
        # drop empty optionals to keep the JSON tidy
        for k in ("url", "dateDisplay", "originalTitle", "originalText", "notes",
                  "furtherBibliography", "commentOnText"):
            if not rec[k]:
                rec.pop(k)
        records.append(rec)

    # ---- images ----
    matched_files = set()
    canon_cache = {}   # source path -> chosen canonical dest filename (stable per file)
    used_dest = set()  # canonical names already taken (collision guard)
    unmatched_files = []
    flags = []  # naming inconsistencies / images-without-records
    carried = 0
    if args.carry_images:
        # Data-only drop: no new image folder shipped. The images already sitting
        # in public/img/recipeImages were matched on a previous run and named after
        # record ids; re-attach them by reading the URLs out of the existing
        # data.json (keyed by the unchanged id / ABBR.2). IMG_DIR is left untouched.
        prev_images = {}
        if OUT_JSON.exists():
            for rec0 in json.loads(OUT_JSON.read_text(encoding="utf-8")):
                if rec0.get("images"):
                    prev_images[rec0["id"]] = rec0["images"]
        for rec in records:
            imgs = prev_images.get(rec["id"])
            if imgs:
                rec["images"] = list(imgs)
                carried += 1
    elif not args.no_images and args.images.exists():
        if IMG_DIR.exists():
            shutil.rmtree(IMG_DIR)
        IMG_DIR.mkdir(parents=True, exist_ok=True)
        files = collect_images(args.images)

        # Year used for image matching comes from the record's id (the ABBR.2
        # convention that filenames also follow), NOT the date column — those can
        # disagree (e.g. Zobi: id 1841.Zob.01 but date 1853).
        def id_year(rec):
            m = re.match(r"(\d{3,4})", _norm(rec["id"]))
            return int(m.group(1)) if m else None

        # build record indexes
        valid_abbrs = set()
        by_full, by_ni, by_abbr = {}, {}, {}
        years_per_abbr = {}
        for rec in records:
            ks = keys_from_record(rec["id"])
            if not ks:
                continue
            a, idx, letter = ks[0]
            valid_abbrs.add(a)
            by_full.setdefault((a, idx, letter), []).append(rec)
            by_ni.setdefault((a, idx), []).append(rec)
            by_abbr.setdefault(a, []).append(rec)
            if id_year(rec):
                years_per_abbr.setdefault(a, set()).add(id_year(rec))

        def folder_year(folder):
            m = re.search(r"(\d{3,4})", folder)
            return int(m.group(1)) if m else None

        def records_for_folder(folder):
            """Records owned by a folder: same year, narrowed by author/site name
            when several distinct sources share that year (e.g. 1703 Moxon vs Neves)."""
            fy = folder_year(folder)
            recs = [r for r in records if id_year(r) == fy]
            authors = {r["authorLastName"].lower() for r in recs}
            if len(authors) > 1:
                fn = _norm(folder)
                narrowed = [r for r in recs
                            if (r["authorLastName"] and _norm(r["authorLastName"]) in fn)
                            or (r.get("site") and _norm(r["site"]) in fn)]
                if narrowed:
                    return narrowed
            return recs

        def canonical_dest(p, rec):
            """One stable <YEAR>.<Abbr>.<tail> name per source file, derived from
            the matched record's id. Cached so a shared image keeps one name, and
            de-duplicated against names already taken."""
            if p in canon_cache:
                return canon_cache[p]
            ext = p.suffix.lower()
            m = re.match(r"(\d{3,4})\.([^.]+)", rec["id"])
            year = m.group(1) if m else ""
            abbr = re.sub(r"[^A-Za-z0-9]", "", strip_diacritics(m.group(2))) if m else "x"
            tail = make_tail(p.stem, year, abbr)
            base = f"{year}.{abbr}.{tail}" if tail else f"{year}.{abbr}"
            base = re.sub(r"\.+", ".", base).strip(".")
            dest = f"{base}{ext}"
            n = 2
            while dest in used_dest:
                dest = f"{base}.{n}{ext}"
                n += 1
            used_dest.add(dest)
            canon_cache[p] = dest
            return dest

        def attach(rec, p):
            dest = canonical_dest(p, rec)
            shutil.copy2(p, IMG_DIR / dest)
            url = f"{IMG_URL_PREFIX}/{dest}"
            if url not in rec["images"]:
                rec["images"].append(url)
            matched_files.add(p)

        def by_year_filter(cands, fyear):
            """If this abbr is reused across years, keep only the matching year."""
            if fyear and cands:
                a = keys_from_record(cands[0]["id"])[0][0]
                if len(years_per_abbr.get(a, set())) > 1:
                    same = [c for c in cands if id_year(c) == fyear]
                    return same or cands
            return cands

        for p in files:
            info = parse_filename(p.name, valid_abbrs)
            fyear = info["year"] or folder_year(p.parent.name)

            if info["keys"]:  # indexed image -> exact recipe(s)
                for (a, idx, letter) in info["keys"]:
                    cands = by_full.get((a, idx, letter)) or by_ni.get((a, idx)) or []
                    # an explicit file year that no candidate shares means the image
                    # belongs to a different (missing) source — don't mis-attach it.
                    if fyear and cands and not any(id_year(c) == fyear for c in cands):
                        flags.append(f"image without a recipe: {p.relative_to(args.images)} "
                                     f"(no {a!r} record for year {fyear})")
                        continue
                    for rec in (c for c in cands if not fyear or id_year(c) == fyear):
                        attach(rec, p)
            elif info["abbr"]:  # source-level but author known (lettered plate, cover)
                for rec in by_year_filter(by_abbr.get(info["abbr"], []), fyear):
                    attach(rec, p)
            elif info["foreign_indexed"]:  # e.g. 'Rai_1399_*' but records say 'Yor'
                flags.append(f"naming inconsistency, NOT mapped: {p.relative_to(args.images)} "
                             f"(file code '{info['abbr_like']}' has no matching record)")
            else:  # cover / title page / page scan with no author code -> use folder
                recs = records_for_folder(p.parent.name)
                if not recs:
                    flags.append(f"image without a recipe: {p.relative_to(args.images)} "
                                 f"(no record for year {fyear})")
                    continue
                if len(recs) == 1:  # single-recipe source -> everything belongs to it
                    attach(recs[0], p)
                elif info["pages"]:  # page scan -> recipe(s) on that page
                    sel = [r for r in recs if info["pages"] & page_numbers(r["page"])]
                    for rec in sel:
                        attach(rec, p)
                else:  # generic cover/title across a multi-recipe source -> all of them
                    for rec in recs:
                        attach(rec, p)

        unmatched_files = [p for p in files if p not in matched_files]

    # sort + dedupe image lists; drop empty
    for rec in records:
        if rec["images"]:
            rec["images"] = sorted(dict.fromkeys(rec["images"]))
        else:
            rec.pop("images")

    OUT_JSON.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    # ---- report ----
    print(f"✓ wrote {len(records)} records -> {OUT_JSON.relative_to(PROJECT)}")
    with_img = sum(1 for r in records if r.get("images"))
    print(f"  records with images : {with_img}")
    if args.carry_images:
        print(f"  carried from data.json: {carried} records (public/img untouched)")
        no_img = [r["id"] for r in records if not r.get("images")]
        print(f"  records WITHOUT images: {len(no_img)}")
    else:
        print(f"  matched image files : {len(matched_files)}")
        print(f"  unmatched images    : {len(unmatched_files)}")
        for p in unmatched_files:
            print(f"      · {p.relative_to(args.images)}")
    if flags:
        print(f"  ⚠ flags ({len(flags)}) — need a human decision:")
        for f in flags:
            print(f"      · {f}")
    if ranges:
        print(f"  year ranges/odd ({len(ranges)}): " +
              ", ".join(f"{a}:{d}" for a, d in ranges[:8]) + (" ..." if len(ranges) > 8 else ""))
    # ingredient sanity sample
    print("  ingredient parse sample:")
    shown = 0
    for rec in records:
        if rec["ingredients"] and shown < 6:
            ing = rec["ingredients"]
            print(f"      {rec['id']}: " + " | ".join(
                f"[{i['amount']}|{i['unit']}|{i['ingredient']}]" for i in ing))
            shown += 1


if __name__ == "__main__":
    main()
